"""
This module performs post-processing operations after the closing yields workflow has successfully completed.
It is designed to run as an add-on to the main project and does not affect email reporting.
"""

import pandas as pd
import logging
from datetime import datetime, timedelta
from pathlib import Path
import os
import sys
import openpyxl  # For Excel file manipulation
from decimal import Decimal  # Import Decimal for precise decimal handling
from openpyxl.styles import Font, PatternFill

# Get the correct paths for imports
current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.append(current_dir)

from config import Config
from workflow_result import WorkflowResult

# Set up logging - separate from the main workflow logs
logger = logging.getLogger('post_processing')
logger.setLevel(logging.INFO)

# Create a file handler
log_file = Config.get_logs_path() / f'post_processing_{datetime.now().strftime("%Y%m%d")}.log'
file_handler = logging.FileHandler(log_file)
file_handler.setLevel(logging.INFO)

# Create a formatter
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)

# Add the handler to the logger
logger.addHandler(file_handler)

class PostProcessor:
    """
    Post-processor for additional operations after the closing yields workflow.
    This class handles any additional processing that needs to happen after 
    the closing yields workflow has completed successfully.
    """
    
    def __init__(self, closing_yields_data: pd.DataFrame):
        """
        Initialize the post-processor with the closing yields data.
        
        Args:
            closing_yields_data: DataFrame containing the processed closing yields
        """
        self.closing_yields_data = closing_yields_data
        logger.info("Initialized PostProcessor with closing yields data")
        self.excel_path = Path(Config.BASE_DIR) / "Bond Price Calculator.xlsx"
        
        # Detect if we're running on a weekend
        self.is_weekend = datetime.now().weekday() >= 5
        if self.is_weekend:
            logger.info("Weekend processing detected")
        else:
            logger.info("Weekday processing detected")
    
    def find_last_data_row(self, sheet):
        """
        Find the last row with data in the Excel sheet.
        
        Args:
            sheet: Excel worksheet to analyze
            
        Returns:
            int: The last row number containing data
        """
        # Start from the top and find the last row with data in column A (date column)
        last_row = 2  # Minimum is row 2 (header is row 1)
        
        for row in range(3, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value is not None:
                last_row = row
        
        logger.info(f"Found last data row at position {last_row}")
        return last_row
    
    def process_data(self) -> pd.DataFrame:
        """
        Update the Bond Price Calculator Excel file with the latest closing yields.
        
        This method:
        1. Reads the "Input" sheet from the Bond Price Calculator Excel file
        2. Gets the security names from the second row (index 1)
        3. Creates a new row with today's date and the closing yields from our data
        4. Adds this new row to the Excel sheet
        5. Saves the updated Excel file
        
        Returns:
            DataFrame containing the post-processed data (same as input data)
        """
        try:
            logger.info("Starting post-processing operations")
            
            # Verify Excel file exists
            if not self.excel_path.exists():
                logger.error(f"Excel file not found at {self.excel_path}")
                raise FileNotFoundError(f"Bond Price Calculator Excel file not found at {self.excel_path}")
            
            logger.info(f"Reading Excel file from {self.excel_path}")
            
            # Load the workbook and select the Input sheet
            try:
                workbook = openpyxl.load_workbook(self.excel_path)
                input_sheet = workbook["Input"]
                logger.info("Successfully opened 'Input' sheet")
            except Exception as e:
                logger.error(f"Error opening Excel sheet: {str(e)}")
                raise
            
            # Get security names from the second row (index 1)
            securities = {}
            for col in range(2, input_sheet.max_column + 1):  # Start from second column
                cell_value = input_sheet.cell(row=2, column=col).value
                if cell_value:
                    securities[str(cell_value).strip()] = col
            
            logger.info(f"Found {len(securities)} securities in the Excel sheet")
            
            # Map securities to closing yields
            securities_to_yields = {}
            for _, row in self.closing_yields_data.iterrows():
                if pd.notna(row['Security']) and pd.notna(row['Closing Yield']):
                    # Convert to Decimal for precise representation
                    try:
                        securities_to_yields[row['Security']] = Decimal(str(row['Closing Yield']))
                    except:
                        # Fallback to original value if conversion fails
                        securities_to_yields[row['Security']] = row['Closing Yield']
            
            logger.info(f"Mapped {len(securities_to_yields)} securities to their closing yields")
            
            # Find the last row with data
            last_row = self.find_last_data_row(input_sheet)
            
            # Find the template row to copy formatting from
            # Using last_row as our formatting template
            template_row = last_row
            logger.info(f"Using row {template_row} as formatting template")
            
            # Next row is immediately after the last data row
            next_row = last_row + 1
            logger.info(f"Adding new data at row {next_row}")
            
            # Get today's date - Using today's date, but the formatting will be adjusted
            today_date = datetime.now().date()
            
            # Write today's date in the first column and copy formatting from template
            new_date_cell = input_sheet.cell(row=next_row, column=1)
            template_date_cell = input_sheet.cell(row=template_row, column=1)
            
            # Set the value (date only, no timestamp)
            new_date_cell.value = today_date
            
            # Preserve the existing custom date format from the template cell
            if template_date_cell.number_format:
                new_date_cell.number_format = template_date_cell.number_format
                logger.info(f"Preserved existing date format: {template_date_cell.number_format}")
            
            # Copy other formatting properties from template cell
            self._copy_cell_format(template_date_cell, new_date_cell)
            
            # Match securities and write closing yields
            securities_written = 0
            for security, col in securities.items():
                if security in securities_to_yields:
                    # Get the template cell to copy formatting from 
                    template_cell = input_sheet.cell(row=template_row, column=col)
                    
                    # Get the cell we want to write to
                    new_cell = input_sheet.cell(row=next_row, column=col)
                    
                    # Get the value as Decimal to preserve precision
                    yield_value = securities_to_yields[security]
                    
                    # Set the value in the Excel cell
                    # Preserve the exact value (don't convert to string and back)
                    if isinstance(yield_value, Decimal):
                        new_cell.value = float(yield_value)
                    else:
                        new_cell.value = yield_value
                    
                    # Preserve existing number format by copying formatting
                    self._copy_cell_format(template_cell, new_cell)
                    
                    securities_written += 1
                else:
                    logger.warning(f"Security {security} not found in closing yields data")
            
            logger.info(f"Added closing yields for {securities_written} securities")
            
            # Check for any securities in our data that were not found in the Excel sheet
            missing_securities = [sec for sec in securities_to_yields.keys() if sec not in securities]
            if missing_securities:
                logger.warning(f"These securities were in our data but not in Excel: {missing_securities}")
            
            # Apply formula extension to the GC sheet
            logger.info("Now extending formulas in the GC sheet")
            gc_success = self.extend_gc_sheet_formulas(workbook)
            
            # Save the updated workbook
            try:
                workbook.save(self.excel_path)
                logger.info(f"Successfully saved updated Excel file to {self.excel_path}")
            except Exception as e:
                logger.error(f"Error saving Excel file: {str(e)}")
                raise
            
            logger.info("Post-processing operations completed successfully")
            
            # Add timestamp and other columns
            self.closing_yields_data['Processing_Timestamp'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # You can add additional calculations as needed
            # For example, calculate yield deviation from previous day
            self.closing_yields_data['Yield_Deviation'] = 0.0  # Placeholder, you can implement actual calculations
            
            # Return the original DataFrame unchanged
            return self.closing_yields_data.copy()
            
        except Exception as e:
            logger.error(f"Error during post-processing: {str(e)}")
            raise
    
    def _copy_cell_format(self, source_cell, target_cell):
        """
        Copy formatting from source cell to target cell.
        This includes font, border, alignment, fill, etc.
        
        Args:
            source_cell: The cell to copy formatting from
            target_cell: The cell to apply formatting to
        """
        try:
            # Set font size to 12 for all cells directly instead of copying
            target_cell.font = Font(size=12)
            
            # For other properties, try to copy but don't throw errors if it fails
            try:
                # Copy border
                if source_cell.border:
                    target_cell.border = source_cell.border
            except Exception as e:
                logger.debug(f"Error copying border: {str(e)}")
            
            try:
                # Copy alignment
                if source_cell.alignment:
                    target_cell.alignment = source_cell.alignment
            except Exception as e:
                logger.debug(f"Error copying alignment: {str(e)}")
            
            try:
                # Copy fill
                if source_cell.fill:
                    target_cell.fill = source_cell.fill
            except Exception as e:
                logger.debug(f"Error copying fill: {str(e)}")
            
            try:
                # Copy protection
                if source_cell.protection:
                    target_cell.protection = source_cell.protection
            except Exception as e:
                logger.debug(f"Error copying protection: {str(e)}")
            
            try:
                # Always copy number format if available
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
            except Exception as e:
                logger.debug(f"Error copying number format: {str(e)}")
            
            logger.debug(f"Applied formatting to cell {target_cell.coordinate} with font size 12")
        except Exception as e:
            logger.warning(f"Error applying cell format: {str(e)}")
            # Don't raise the exception - it's not critical if formatting fails
    
    def extend_gc_sheet_formulas(self, workbook):
        """
        Find the GC sheet, identify the last row with data, and copy it to the next row.
        Let Excel handle formula adjustments, and update the SETTLE_DATE to the next day.
        
        Args:
            workbook: The openpyxl workbook object
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Check if GC sheet exists
            if "GC" not in workbook.sheetnames:
                logger.warning("GC sheet not found in workbook")
                return False
                
            gc_sheet = workbook["GC"]
            logger.info("Successfully opened 'GC' sheet")
            
            # Find the last row with data (starts from column A)
            last_row = 2  # Default to row 2 if no data found
            for row in range(2, gc_sheet.max_row + 1):
                if gc_sheet.cell(row=row, column=1).value is not None:
                    last_row = row
            
            logger.info(f"Found last data row in GC sheet at row {last_row}")
            
            # Target row where we want to add data
            target_row = last_row + 1
            
            # Find the SETTLE_DATE column - search by column header
            settle_date_col = None
            for col in range(1, gc_sheet.max_column + 1):
                header_cell = gc_sheet.cell(row=1, column=col)
                if header_cell.value and isinstance(header_cell.value, str) and "SETTLE" in header_cell.value.upper():
                    settle_date_col = col
                    logger.info(f"Found SETTLE_DATE column at column {settle_date_col} (header: {header_cell.value})")
                    break
            
            if settle_date_col is None:
                logger.warning("Could not find SETTLE_DATE column in GC sheet, will try to look for date values instead")
                # Alternative approach: look for a column that contains date values
                for col in range(1, gc_sheet.max_column + 1):
                    cell_value = gc_sheet.cell(row=last_row, column=col).value
                    if isinstance(cell_value, datetime):
                        settle_date_col = col
                        logger.info(f"Found date column at column {settle_date_col} based on data type")
                        break
            
            # Now copy all cells from the last row to the target row
            # We'll do this in two steps:
            # 1. First, copy all values, formulas, and formats
            # 2. Then, specifically update the SETTLE_DATE if found
            
            # Step 1: Copy all cells
            for col in range(1, gc_sheet.max_column + 1):
                source_cell = gc_sheet.cell(row=last_row, column=col)
                target_cell = gc_sheet.cell(row=target_row, column=col)
                
                # Get formula or value from source cell
                if source_cell.value is not None:
                    if isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                        # This is a formula, we'll adjust it
                        formula = source_cell.value
                        
                        # Replace row references
                        # Example: If formula contains A10, B10, etc. and we're copying from row 10 to 11,
                        # we need to change A10, B10 to A11, B11
                        import re
                        
                        # Find all cell references in the formula
                        # Example pattern: A10, B12, AA123, etc.
                        cell_ref_pattern = r'([A-Z]+)(' + str(last_row) + r')'
                        refs = re.findall(cell_ref_pattern, formula)
                        
                        # Create new formula with adjusted references
                        new_formula = formula
                        for col_ref, row_ref in refs:
                            old_ref = f"{col_ref}{row_ref}"
                            new_ref = f"{col_ref}{target_row}"
                            new_formula = new_formula.replace(old_ref, new_ref)
                        
                        target_cell.value = new_formula
                        logger.debug(f"Adjusted formula from {formula} to {new_formula}")
                    else:
                        # This is a regular value, just copy it
                        target_cell.value = source_cell.value
                
                # Copy number format if available
                if source_cell.number_format:
                    try:
                        target_cell.number_format = source_cell.number_format
                    except Exception as e:
                        logger.debug(f"Could not copy number format: {str(e)}")
                
                # Apply other formatting (with font size 12)
                self._copy_cell_format(source_cell, target_cell)
            
            # Step 2: If we found a SETTLE_DATE column, update the date
            if settle_date_col:
                logger.info(f"Updating SETTLE_DATE in column {settle_date_col}")
                source_cell = gc_sheet.cell(row=last_row, column=settle_date_col)
                target_cell = gc_sheet.cell(row=target_row, column=settle_date_col)
                
                # Get the current date value
                current_date = source_cell.value
                
                # Try to convert to datetime if it's not already
                if not isinstance(current_date, datetime):
                    try:
                        if isinstance(current_date, str):
                            current_date = pd.to_datetime(current_date)
                        else:
                            # If we can't parse it, log a warning and skip date update
                            logger.warning(f"SETTLE_DATE value '{current_date}' is not a recognizable date format")
                            # We'll continue with other operations
                    except Exception as e:
                        logger.warning(f"Error parsing SETTLE_DATE: {str(e)}")
                        # We'll continue with other operations
                
                # If we have a valid date, increment it by one day
                if isinstance(current_date, datetime):
                    new_date = current_date + pd.Timedelta(days=1)
                    
                    # Make sure we only have the date part, no time
                    if hasattr(new_date, 'date'):
                        new_date = new_date.date()
                    
                    # Set the new date in the target cell
                    target_cell.value = new_date
                    
                    # Preserve existing date format - let Excel handle display conversion
                    if source_cell.number_format:
                        target_cell.number_format = source_cell.number_format
                    
                    logger.info(f"Updated SETTLE_DATE from {current_date} to {new_date}")
                    
                    # Reapply other formatting to ensure it shows correctly
                    self._copy_cell_format(source_cell, target_cell)
            
            logger.info(f"Successfully copied last row to row {target_row} in GC sheet with adjusted formulas")
            
            return True
            
        except Exception as e:
            logger.error(f"Error extending GC sheet: {str(e)}")
            return False
    
    def save_results(self, df: pd.DataFrame) -> Path:
        """
        Save the post-processed data to a CSV file.
        
        Args:
            df: DataFrame containing the post-processed data
            
        Returns:
            Path to the saved CSV file
        """
        try:
            # Save to today's output directory with a different name
            output_file = Config.get_output_path() / f'post_processed_data_{datetime.now().strftime("%Y%m%d")}.csv'
            
            # Save to CSV
            df.to_csv(output_file, index=False)
            logger.info(f"Successfully saved post-processed data to {output_file}")
            
            return output_file
            
        except Exception as e:
            logger.error(f"Error saving post-processed data: {str(e)}")
            raise

    def process_weekend_update(self) -> bool:
        """
        Simplified weekend processing method.
        Instead of using closing yields data, this method simply:
        1. Opens the existing Excel file
        2. Finds the last row with data
        3. Copies that row to a new row
        4. Updates the date column to today's date (Saturday or Sunday)
        5. Updates the GC sheet with extended formulas
        
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            logger.info(f"Starting simplified weekend processing using Excel file at {self.excel_path}")
            
            if not self.excel_path.exists():
                logger.error(f"Excel file not found at {self.excel_path}")
                return False
                
            # Load the workbook
            workbook = openpyxl.load_workbook(self.excel_path)
            
            # Check if Input sheet exists
            if "Input" not in workbook.sheetnames:
                logger.error("Could not find 'Input' sheet in Excel file")
                return False
                
            input_sheet = workbook["Input"]
            
            # Find the last row with data
            last_row = self.find_last_data_row(input_sheet)
            target_row = last_row + 1
            
            logger.info(f"Found last data row at {last_row}, will insert new row at {target_row}")
            
            # Copy all cells from last row to target row
            for col in range(1, input_sheet.max_column + 1):
                source_cell = input_sheet.cell(row=last_row, column=col)
                target_cell = input_sheet.cell(row=target_row, column=col)
                
                # Copy value
                target_cell.value = source_cell.value
                
                # Copy formatting
                self._copy_cell_format(source_cell, target_cell)
            
            # Update the date column (column 1) to today's date (Saturday or Sunday)
            last_date = input_sheet.cell(row=last_row, column=1).value
            if isinstance(last_date, datetime):
                # Use today's date (which will be a Saturday or Sunday in weekend mode)
                # Use only the date part, no time component
                today_date = datetime.now().date()
                
                logger.info(f"Updating date from {last_date} to today's date: {today_date}")
                input_sheet.cell(row=target_row, column=1).value = today_date
                
                # Preserve the existing date format from the template
                # Excel will handle the display conversion automatically
                if input_sheet.cell(row=last_row, column=1).number_format:
                    input_sheet.cell(row=target_row, column=1).number_format = input_sheet.cell(row=last_row, column=1).number_format
            else:
                logger.warning(f"Date cell does not contain a valid date: {last_date}")
            
            # Step 2: Update the GC sheet with extended formulas
            logger.info("Now processing GC sheet to extend formulas...")
            
            # Call the method to extend formulas in the GC sheet
            gc_success = self.extend_gc_sheet_formulas(workbook)
            
            if not gc_success:
                logger.warning("Could not extend GC sheet formulas, but Input sheet was updated successfully.")
            else:
                logger.info("Successfully extended GC sheet formulas")
            
            # Save the workbook
            workbook.save(self.excel_path)
            logger.info(f"Successfully saved updated Excel file to {self.excel_path}")
            
            return True
            
        except Exception as e:
            logger.error(f"Error in weekend processing: {str(e)}")
            return False

def run_post_processing_workflow(closing_yields_data: pd.DataFrame = None, is_weekend_mode: bool = False) -> WorkflowResult:
    """
    Run the post-processing workflow.
    
    In weekend mode:
    - Uses simplified processing that directly copies the last row in the Excel file
    - Does not require closing yields data
    
    In weekday mode:
    - Uses the full processing with closing yields data
    - Requires closing yields data to be provided
    
    Args:
        closing_yields_data: DataFrame containing the processed closing yields (required for weekday mode)
        is_weekend_mode: Flag indicating if running in weekend mode with simplified processing
        
    Returns:
        WorkflowResult containing success status and processed data
    """
    try:
        if is_weekend_mode:
            logger.info("Starting post-processing workflow in WEEKEND MODE with simplified processing")
            
            # For weekend mode, we don't need closing yields data
            processor = PostProcessor(pd.DataFrame() if closing_yields_data is None else closing_yields_data)
            
            # Use the simplified weekend processing method
            success = processor.process_weekend_update()
            
            if success:
                logger.info("Successfully completed weekend post-processing")
                return WorkflowResult(success=True, data=pd.DataFrame())
            else:
                error_msg = "Weekend post-processing failed"
                logger.error(error_msg)
                return WorkflowResult(success=False, error=error_msg)
        else:
            # Weekday mode requires closing yields data
            if closing_yields_data is None:
                error_msg = "Closing yields data is required for weekday processing"
                logger.error(error_msg)
                return WorkflowResult(success=False, error=error_msg)
                
            logger.info("Starting post-processing workflow")
            
            # Initialize processor with closing yields data
            processor = PostProcessor(closing_yields_data)
            
            # Process the data
            results_df = processor.process_data()
            
            # Save results
            output_file = processor.save_results(results_df)
            
            logger.info("Successfully completed post-processing workflow")
            
            return WorkflowResult(success=True, data=results_df)
        
    except Exception as e:
        error_msg = f"Error in post-processing workflow: {str(e)}"
        logger.error(error_msg)
        return WorkflowResult(success=False, error=error_msg)

# This section only runs if you execute this file directly (not when imported)
if __name__ == "__main__":
    try:
        # Check if today is a weekend
        is_weekend = datetime.now().weekday() >= 5
        
        if is_weekend:
            print("Weekend detected - using simplified Excel update process")
            
            # Use simplified weekend processing that doesn't require closing yields data
            result = run_post_processing_workflow(is_weekend_mode=True)
            
            if result.success:
                print("Weekend post-processing completed successfully")
                print("The last row in the Excel file was copied and the date was updated")
            else:
                print(f"Weekend post-processing failed: {result.error}")
        else:
            print("Weekday detected - looking for closing yields data")
            
            # On weekdays, use today's file if it exists
            today_str = datetime.now().strftime("%Y%m%d")
            closing_yields_file = Config.get_output_path() / f'closing_yields_{today_str}.csv'
            
            if not closing_yields_file.exists():
                print(f"Error: Today's closing yields file not found at {closing_yields_file}")
                print("Please run the closing yields workflow first")
                exit(1)
            
            # Load the closing yields data
            print(f"Loading data from: {closing_yields_file}")
            closing_yields_data = pd.read_csv(closing_yields_file)
            
            # Run the post-processing workflow
            result = run_post_processing_workflow(closing_yields_data, is_weekend_mode=False)
            
            if result.success:
                print("Post-processing completed successfully")
                print("Post-processed data preview:")
                print(result.data.head())
            else:
                print(f"Post-processing failed: {result.error}")
            
    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        print(traceback.format_exc()) 